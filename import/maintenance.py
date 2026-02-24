import openpyxl
from db import Checks

wb = openpyxl.load_workbook('C:\session_chmpion\sesseion main\session 1\Приложение Ресурсы\Ресурсы\Import/maintenance.xlsx')
ws = wb['Sheet1']
for i in range(2,102):
    Checks.create(
        date = ws[f"A{i}"].value,
        issues_found = ws[f"B{i}"].value,
        vending_machine_id = ws[f"C{i}"].value,
        full_name = ws[f"D{i}"].value,
        work_description = ws[f"E{i}"].value
    )
